import os
import re
import time
import urllib.parse
import requests
import smtplib
import ssl
from email.message import EmailMessage
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font

# ============== ENV (GitHub Secrets) ==============
SERPAPI_KEY = os.getenv("SERPAPI_KEY")

EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")  # Gmail App Password

# Put 2 emails in GitHub secret EMAIL_RECEIVER like:
# first@gmail.com, second@gmail.com
EMAIL_RECEIVERS = [e.strip() for e in (os.getenv("EMAIL_RECEIVER") or "").split(",") if e.strip()]
# ================================================

LOCATION = "United States"

ROLE_KEYWORDS = [
    "Quality Assurance Supervisor",
    "Quality Assurance Manager",
    "QA Supervisor",
    "QA Manager",
    "Quality Supervisor",
    "Quality Manager",
    "FSQ Manager",
    "FSQ Supervisor",
    "FSQ Specialist",
    "FSQA Manager",
    "FSQA Supervisor",
    "FSQA Specialist",
    "Food Safety Manager",
    "Food Safety Supervisor",
    "Quality Specialist",
    "Quality Lead",
]

FOOD_HINTS = [
    "food", "food manufacturing", "food processing", "meat", "dairy", "bakery", "beverage",
    "plant", "production", "warehouse", "HACCP", "SQF", "FSQA", "GMP", "sanitation"
]


def validate_env():
    if not SERPAPI_KEY:
        raise ValueError("SERPAPI_KEY missing (GitHub Secret).")
    if not EMAIL_SENDER or not EMAIL_PASSWORD or not EMAIL_RECEIVERS:
        raise ValueError("EMAIL_SENDER / EMAIL_PASSWORD / EMAIL_RECEIVER missing (GitHub Secrets).")


# ---------------- SerpAPI calls with retry/backoff ----------------
def serpapi_google_jobs(query: str, location: str, num: int = 50) -> List[Dict[str, Any]]:
    params = {
        "engine": "google_jobs",
        "q": query,
        "location": location,
        "api_key": SERPAPI_KEY,
        "num": num,
    }

    retry_statuses = {429, 502, 503, 504}
    max_attempts = 5

    for attempt in range(1, max_attempts + 1):
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=30)

            if r.status_code in retry_statuses:
                time.sleep(2 ** attempt)
                continue

            r.raise_for_status()
            return r.json().get("jobs_results", []) or []

        except requests.RequestException:
            time.sleep(2 ** attempt)

    return []


def serpapi_google_jobs_listing(job_id: str) -> Dict[str, Any]:
    if not job_id:
        return {}

    params = {"engine": "google_jobs_listing", "job_id": job_id, "api_key": SERPAPI_KEY}
    retry_statuses = {429, 502, 503, 504}
    max_attempts = 4

    for attempt in range(1, max_attempts + 1):
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=30)

            if r.status_code in retry_statuses:
                time.sleep(2 ** attempt)
                continue

            if r.status_code != 200:
                return {}

            return r.json() or {}

        except requests.RequestException:
            time.sleep(2 ** attempt)

    return {}


# ---------------- Helpers ----------------
def safe_apply_link(job: Dict[str, Any]) -> str:
    links = job.get("related_links") or []
    if isinstance(links, list) and links:
        return links[0].get("link") or "N/A"
    return "N/A"


def safe_source_link(job: Dict[str, Any]) -> str:
    links = job.get("related_links") or []
    if isinstance(links, list) and len(links) >= 2:
        return links[1].get("link") or "N/A"
    if isinstance(links, list) and len(links) == 1:
        return links[0].get("link") or "N/A"
    return "N/A"


def safe_apply_link_from_details(details: Dict[str, Any]) -> str:
    apply_options = details.get("apply_options") or []
    if isinstance(apply_options, list) and apply_options:
        return apply_options[0].get("link") or "N/A"
    return "N/A"


def safe_source_link_from_details(details: Dict[str, Any]) -> str:
    apply_options = details.get("apply_options") or []
    if isinstance(apply_options, list) and len(apply_options) >= 2:
        return apply_options[1].get("link") or "N/A"
    if isinstance(apply_options, list) and len(apply_options) == 1:
        return apply_options[0].get("link") or "N/A"
    return "N/A"


def safe_pay(job: Dict[str, Any]) -> str:
    de = job.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("salary"):
        return str(de["salary"])

    ext = job.get("extensions") or []
    if isinstance(ext, list):
        for item in ext:
            if isinstance(item, str) and ("$" in item or "hour" in item.lower() or "year" in item.lower()):
                return item
    return "N/A"


def safe_pay_from_details(details: Dict[str, Any]) -> str:
    de = details.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("salary"):
        return str(de["salary"])
    return "N/A"


def safe_time_posted(job: Dict[str, Any]) -> str:
    de = job.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("posted_at"):
        return str(de["posted_at"])

    ext = job.get("extensions") or []
    if isinstance(ext, list):
        for item in ext:
            if isinstance(item, str) and (
                "ago" in item.lower() or "today" in item.lower() or
                "yesterday" in item.lower() or "posted" in item.lower()
            ):
                return item
    return "N/A"


def safe_time_posted_from_details(details: Dict[str, Any]) -> str:
    de = details.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("posted_at"):
        return str(de["posted_at"])
    return "N/A"


def posted_days(time_posted: str) -> int:
    if not time_posted or time_posted == "N/A":
        return 999

    s = time_posted.strip().lower()
    if "just posted" in s or "today" in s:
        return 0
    if "yesterday" in s:
        return 1

    if re.search(r"\d+\s+hour", s):
        return 0

    m = re.search(r"(\d+)\s+day", s)
    if m:
        return int(m.group(1))

    m = re.search(r"(\d+)\s+week", s)
    if m:
        return int(m.group(1)) * 7

    return 999


def looks_food_industry(job: Dict[str, Any]) -> bool:
    text = " ".join([
        str(job.get("title") or ""),
        str(job.get("company_name") or ""),
        str(job.get("description") or ""),
    ]).lower()
    return any(h.lower() in text for h in FOOD_HINTS)


def company_careers_search_link(company_name: str) -> str:
    if not company_name or company_name == "N/A":
        return "N/A"
    q = f"{company_name} careers"
    return "https://www.google.com/search?q=" + urllib.parse.quote_plus(q)


def normalize_row(job: Dict[str, Any]) -> Dict[str, str]:
    job_id = job.get("job_id") or "N/A"

    title = job.get("title") or "N/A"
    company = job.get("company_name") or "N/A"
    location = job.get("location") or "N/A"
    source = job.get("via") or "Unknown"

    pay = safe_pay(job)
    time_posted = safe_time_posted(job)
    apply_link = safe_apply_link(job)
    source_link = safe_source_link(job)

    if job_id != "N/A" and (pay == "N/A" or time_posted == "N/A" or apply_link == "N/A" or source_link == "N/A"):
        details = serpapi_google_jobs_listing(job_id)
        if details:
            if pay == "N/A":
                pay = safe_pay_from_details(details) or pay
            if time_posted == "N/A":
                time_posted = safe_time_posted_from_details(details) or time_posted
            if apply_link == "N/A":
                apply_link = safe_apply_link_from_details(details) or apply_link
            if source_link == "N/A":
                source_link = safe_source_link_from_details(details) or source_link
            if source == "Unknown":
                source = details.get("via") or source

    return {
        "job_id": job_id,
        "title": title,
        "company_name": company,
        "pay": pay if pay else "N/A",
        "time_posted": time_posted if time_posted else "N/A",
        "location": location,
        "source": source,
        "apply_link": apply_link if apply_link else "N/A",
        "source_link": source_link if source_link else "N/A",
        "company_careers_link": company_careers_search_link(company),
    }


def build_queries() -> List[str]:
    queries = []
    for role in ROLE_KEYWORDS:
        queries.append(f'"{role}" food')
        queries.append(f'"{role}" food manufacturing')
        queries.append(f'"{role}" food processing')
        queries.append(f'"{role}" HACCP')
        queries.append(f'"{role}" SQF')
        queries.append(f'"{role}" FSQA')
    return queries


def dedupe_by_job_id(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    out = []
    for row in rows:
        key = row.get("job_id") or (row.get("title", "") + "|" + row.get("company_name", "") + "|" + row.get("location", ""))
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def create_excel(rows: List[Dict[str, str]], filename: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "title",
        "company_name",
        "pay",
        "time_posted",
        "location",
        "source",
        "apply_link",
        "source_link",
        "company_careers_link",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h, "N/A") for h in headers])

    link_cols = {
        "apply_link": headers.index("apply_link") + 1,
        "source_link": headers.index("source_link") + 1,
        "company_careers_link": headers.index("company_careers_link") + 1,
    }

    for row_idx in range(2, ws.max_row + 1):
        for _, col_idx in link_cols.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value or "")
            if val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")

    wb.save(filename)
    return filename


def send_email_with_attachment(subject: str, body: str, attachment_path: str):
    msg = EmailMessage()
    msg["From"] = EMAIL_SENDER
    msg["To"] = ", ".join(EMAIL_RECEIVERS)   # ✅ sends to 2 emails
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(attachment_path),
    )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)


def main():
    validate_env()

    all_rows: List[Dict[str, str]] = []

    for q in build_queries():
        jobs = serpapi_google_jobs(q, LOCATION, num=50)
        for job in jobs:
            if looks_food_industry(job):
                all_rows.append(normalize_row(job))

    all_rows = dedupe_by_job_id(all_rows)
    all_rows = [r for r in all_rows if posted_days(r.get("time_posted", "N/A")) <= 7]
    all_rows.sort(key=lambda r: posted_days(r.get("time_posted", "N/A")))

    today = datetime.now().strftime("%Y-%m-%d")
    excel_file = f"food_quality_jobs_{today}.xlsx"
    create_excel(all_rows, excel_file)

    subject = f"Daily Food Quality Jobs Report - {today}"
    body = f"""Hi,

Attached is your daily Food Industry Quality/FSQA job report (last 7 days).
Total jobs found: {len(all_rows)}

Receivers: {", ".join(EMAIL_RECEIVERS)}

Regards,
Job Bot
"""
    send_email_with_attachment(subject, body, excel_file)


if __name__ == "__main__":
    main()
